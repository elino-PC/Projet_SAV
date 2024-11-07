import requests
import os
import json
import datetime
from time import *
from datetime import *
import pandas as pd
pd.options.mode.chained_assignment = None
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.drawing.image import Image
import numpy as np
import pytz
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import ssl
import io
from sklearn.linear_model import LinearRegression

from Core.Installation_Definition import SolarInstallation


class SolarInstallationMC(SolarInstallation):  
    
    MONTH_MAPPING = {
        'Janv': 1, 'Févr': 2, 'Mars': 3, 'Avril': 4, 'Mai': 5, 'Juin': 6,
        'Juil': 7, 'Août': 8, 'Sept': 9, 'Oct': 10, 'Nov': 11, 'Déc': 12
    }
    
    
    def __init__(self, name, report_type, data_path, year):
        super().__init__(name)
        self.type = "meteocontrol"
        
        # Validation du type de rapport
        if report_type not in ['1m', '3m']:
            raise ValueError("Le rapport doit être soit de 1 mois ('1m') soit de 3 mois ('3m')")
        
        self.report_type = report_type
        self.data_path = data_path
        self.year = year
        self.df = None

    
    def load_data(self, min_row=3):
        data_file = load_workbook(self.data_path)
        sheet = data_file.active
        data = sheet.iter_rows(min_row=min_row, values_only=True)
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        df = df.replace('x', 0)
        return df
    
    def convert_dates(self):
        self.df['Date'] = [self.MONTH_MAPPING[month] for month in self.df['Date']]
        self.df['Formatted_Date'] = self.df['Date'].apply(lambda month: f"01/{month:02d}/{self.year}")
        self.df['Formatted_Date'] = pd.to_datetime(self.df['Formatted_Date'], format="%d/%m/%Y")

    def process_columns(self, columns_to_process):
        df_temp = pd.DataFrame()
        df_temp['Date'] = self.df['Formatted_Date']

        for data_column, export_column in columns_to_process:
            matched_columns = [col for col in self.df.columns if data_column.lower() in col.lower()]
            if matched_columns:
                df_temp[export_column] = self.df[matched_columns].sum(axis=1)
            else:
                df_temp[export_column] = 0

        return df_temp


    def get_all_data_month(self, month_type, writer):
        
        """
        Récupère toutes les données pour l'installation en fonction des dates définies.
        """
        
        data = self.load_data()
        self.convert_dates()
        
        columns_to_process = [
        ("Consommation Jirama totale", "Jirama alimentant les charges"),
        ("Consommation charge totale","Consommation totale"),
        ("Énergie active (export)","Export"),
        ("Compteur rendement photovoltaïque", "Production PV totale"),
        ("Production GE totale","Groupe alimentant les charges"),
        ("Fraction renouvelable","Fraction renouvelable"),
        ("Indice de performance énergétique", "Indice de performance énergétique"),
        ("Production spécifique", "Rendement spécifique moyen"),
        ("Ratio de performance", "Performance Ratio"),
        ("Irradiation","Irradiance moyenne")]
        
        df_temp = self.process_columns(columns_to_process)
        
        if 'Export' in df_temp_columns:
            df_temp["Export"] = df_temp["Export"].diff()
            
        df_export = pd.DataFrame()
        df_export['Date'] = df_temp['Date']
        df_export["Groupe"] = df_temp['Groupe alimentant les charges']
        df_export['Jirama'] = df_temp["Jirama alimentant les charges"]
        df_export['PV'] = df_temp["Production PV totale"]
        df_export["Consommation"] = df_temp['Consommation totale']
        df_export["Rendement spécifique moyen"] = df_temp['Rendement spécifique moyen']
        df_export['Irradiance moyenne'] = df_temp['Irradiance moyenne']
        df_export["Export"] = df_temp['Export']
        df_export["Indice de performance énergétique"] = df_temp['Indice de performance énergétique']
        df_export["Performance Ratio"] = df_temp['Performance Ratio']
        
        if month_type == "prev" :
            df_export.to_excel(writer, sheet_name= "data_previous_month", index=False)
        else :
            df_export.to_excel(writer, sheet_name= "data", index=False)
            
            id_target_day_sun = df_export['PV'].idxmax()
            target_day_sun = df_export['Date'][id_target_day_sun].strftime('%d-%m-%Y')
            
            id_target_day_conso = df_export['Consommation'].idxmax()
            target_day_conso = df_export['Date'][id_target_day_conso].strftime('%d-%m-%Y')
            
            data_path_sun = os.path.dirname(data_path) + f"/{target_day_sun}.xlsx"
            data_path_conso = os.path.dirname(data_path) + f"/{target_day_conso}.xlsx"

            self.get_data_site_day(data_path_sun, writer, "sun", target_day_sun)
            self.get_data_site_day(data_path_conso, writer, "conso", target_day_conso)
        
        
        return data_path_sun, data_path_conso, df_export, target_day_conso, target_day_sun
    
    
    def get_data_site_day(self, day_type, target_day, writer):
        
        data = self.load_data()
        self.convert_dates()
        
        df_temp = pd.DataFrame()
        df_temp['Date'] = df['Time']
        
        if 'Export' in df_temp.columns:
            df_temp["Export"] = df_temp["Export"].diff()
        
        df_temp['Date'] = pd.to_datetime(df_temp['Date'])
        df_temp = df_temp.groupby(df_temp['Date'].dt.floor('H')).sum().reset_index()
        df_temp = df_temp.reset_index()

        df_export = pd.DataFrame()
        df_export['Date'] = df_temp['Date']
        df_export['PV directement consommée'] = df_temp["Production PV totale"]
        df_export['PV exporté sur le réseau'] = df_temp["Export"]
        df_export["Groupe alimentant les charges"] = df_temp['Groupe alimentant les charges']
        df_export['Jirama alimentant les charges'] = df_temp["Jirama alimentant les charges"]
        df_export['Irradiance moyenne'] = df_temp['Irradiance moyenne']/12
        df_export["Consommation totale"] = df_temp['Consommation totale']
        
        
        df_export.to_excel(writer, sheet_name=f"data_{day_type}", index=False)
        
        return target_day
    
    
    def get_data_site_12m(self, data_path_12m_prev, data_path_12m, report_date):
        
        list_path = [data_path_12m_prev, data_path_12m]
        df_export = pd.DataFrame()
        
        for path in list_path:
            data = self.load_data(path)
            self.convert_dates()
            df_temp = self.process_columns([
            ("Consommation Jirama totale", "Jirama alimentant les charges"),
            ("Consommation charge totale","Consommation totale"),
            ("Énergie active (export)","Export"),
            ("Compteur rendement photovoltaïque", "Production PV totale"),
            ("Production GE totale","Groupe alimentant les charges"),
            ("Fraction renouvelable","Fraction renouvelable"),
            ("Indice de performance énergétique", "Indice de performance énergétique"),
            ("Production spécifique", "Rendement spécifique moyen"),
            ("Ratio de performance", "Performance Ratio"),
            ("Irradiation","Irradiance moyenne")])
            
            if 'Export' in df_temp.columns:
                df_temp["Export"] = df_temp["Export"].diff()
             
            df_temp = df_temp.fillna(0)
            df_export = pd.DataFrame({
                'Date': df_temp['Date'],
                'Production PV totale': df_temp["Production PV totale"],
                'Consommation totale': df_temp['Consommation totale'],
                'Groupe': df_temp['Groupe alimentant les charges'],
                'Jirama': df_temp["Jirama alimentant les charges"],
                'Export': df_temp['Export'],
                'Irradiance moyenne': df_temp['Irradiance moyenne'] / 30.5
            })
            
            df_export_tot = pd.concat([df_export_tot, df_export], ignore_index=True)

        df_export_tot = df_export_tot[~df_export_tot.iloc[:, 1:].eq(0).all(axis=1)]
        df_export_tot['Date'] = pd.to_datetime(df_export_tot['Date'], format="%d/%m/%Y")
        
        df_export_tot = df_export_tot[df_export_tot['Date'] <= report_date]

        if len(df_export_tot) > 12:
            df_export_tot = df_export_tot.tail(13)
            
        df_export_tot.to_excel(writer, sheet_name="data_12m", index=False)

            

# Fonction pour obtenir les identifiants de site à partir des noms de site
def get_site_list(name_site_list):
    # Charger le fichier Excel contenant les données de site
    df = pd.read_excel("G:/.shortcut-targets-by-id/12F2rxDjhgWoKdVhzLQxaPsc1FL1nWr9H/17- Technique/3 - SAV/3 - Rapports de production/3 - Outil de rapport/Projet_SAV/#Master_Report Generator.xlsm", sheet_name='meteocontrol')

    # Dictionnaire pour stocker les résultats
    site_ids = {}

    # Obtenir l'identifiant pour chaque nom de site
    for name_site in name_site_list:
        # Rechercher le nom du site dans la colonne 'Nom Installation'
        site_row = df[df['Nom Installation'] == name_site]
        if not site_row.empty:
            site_ids[name_site] = site_row.iloc[0]['Nom Installation']
        else:
            site_ids[name_site] = f"Le nom du site '{name_site}' n'a pas été trouvé dans le DataFrame."

    # Retourner le dictionnaire contenant les noms et identifiants des sites
    return site_ids


# Liste des noms de site
#name_site_list = ['Antana Production', 'EPSILON', 'Epsilon - site 2', 'SOCOTA - PHASE 1', 'Actual Textile', 'Menakao']

# Obtenir les identifiants de site pour chaque nom de site
#site_list = get_site_list(name_site_list)

# Afficher les résultats
#for site_name, site_id in site_list.items():
   #print(f"Le nom du site '{site_name}' a pour identifiant de site : {site_id}")
