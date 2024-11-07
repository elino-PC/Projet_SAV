import requests
import http.client
import os
import json
import datetime
from time import * 
from datetime import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import  BarChart, PieChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font 
from openpyxl.drawing.image import Image  
import numpy as np
import pytz
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from sklearn.linear_model import LinearRegression
import shutil

from Core.Installation_Definition import SolarInstallation
from Data_Collection.collectors.PV_Gis import get_irradiance_pv_gis
from Report_Generation.Chart_Generation import create_charts


# le master_generator
def get_recap_values(master_path, installation):
    worksheet_name = installation.type  # Assurez-vous que installation.type est défini
    try:
        # Lister toutes les feuilles disponibles dans le fichier Excel
        xl = pd.ExcelFile(master_path)
        print(f"Feuilles disponibles dans le fichier : {xl.sheet_names}")

        # Vérifier si la feuille existe avant de lire
        if worksheet_name not in xl.sheet_names:
            print(f"La feuille {worksheet_name} n'existe pas dans le fichier.")
            return None

        # Lire la feuille spécifiée
        df = pd.read_excel(master_path, sheet_name=worksheet_name)

        # Filtrer les données correspondant à l'installation
        recap_values_array = df[df.iloc[:, 0] == installation.name]

        if recap_values_array.empty:
            print(f"Aucune donnée trouvée pour l'installation {installation.name} dans la feuille {worksheet_name}")
            return None

        # Extraire les valeurs récapitulatives
        recap_values = recap_values_array.iloc[0].tolist()
        print(f"Valeurs de récapitulatif obtenues pour {installation.name} : {recap_values}")
        return recap_values

    except FileNotFoundError:
        print(f"Le fichier maître à {master_path} est introuvable.")
        return None
    except IndexError:
        print(f"Problème avec l'indexation des valeurs récapitulatives pour {installation.name}.")
        return None
    except Exception as e:
        print(f"Erreur inattendue lors de la récupération des valeurs récapitulatives : {e}")
        return None

    
def create_report_file(installation, template_path, recap_values, result_path):
    # Load the workbook
    workbook = load_workbook(template_path)

    # Save a copy of the workbook that will be the report
    file_name = f'{result_path}/{installation.name}_{recap_values[5].year}_{recap_values[5].month}_{installation.report_type}.xlsx'
        # Copy the template to the destination
    shutil.copy(template_path, file_name)

    # Load the copied workbook
    workbook = load_workbook(file_name)

    # Save the workbook (this step might be redundant but ensures any changes are saved)
    workbook.save(file_name)
    
    # Return name of the report file
    return file_name
    
def fill_aide_rapport (path_rapport, recap_values, installation):
    template_xlsx = load_workbook(path_rapport)
    worksheet = template_xlsx['Aide Rapport']

    # Insallation information
    worksheet['B2'] = installation.name
    worksheet['B7'] = recap_values[7]
    worksheet['B8'] = recap_values[8]
    worksheet['B9'] = recap_values[9]
    worksheet['B10'] = recap_values[10]
    worksheet['D7'] = recap_values[12]
    worksheet['D8'] = recap_values[13]
    worksheet['D9'] = recap_values[14]
    worksheet['D10'] = recap_values[15]
    worksheet['D11'] = recap_values[16]
    worksheet['D4'] = recap_values[17]

    # Useful dates
    # Calculate the start and end dates of the previous month
    s_previous_month = (recap_values[5].replace(day=1) - timedelta(days=1)).replace(day=1)
    e_previous_month = recap_values[5].replace(day=1) - timedelta(days=1)
    
    worksheet['D1'] = s_previous_month.strftime('%d/%m/%Y')
    worksheet['E1'] = e_previous_month.strftime('%d/%m/%Y')
    
    # Calculate the start and end dates of two months prior
    s_previous_2month = (s_previous_month.replace(day=1) - timedelta(days=1)).replace(day=1)
    e_previous_2month = s_previous_month.replace(day=1) - timedelta(days=1)
    
    worksheet['D2'] = s_previous_2month.strftime('%d/%m/%Y')
    worksheet['E2'] = e_previous_2month.strftime('%d/%m/%Y')

    #Calculate the list of dates that will be useful
    if installation.report_type =='1m':
        date_list = pd.date_range(start=recap_values[5], end=recap_values[6]).strftime('%d/%m/%Y').tolist()
        for ii, date in enumerate(date_list):
            cell = worksheet.cell(row=17+ii, column=1)
            cell.value = date
        
        worksheet['B1'] = date_list[0]
        worksheet['C1'] = date_list[-1]
    elif installation.report_type =='3m':
        date_ranges = [
                (recap_values[5], (recap_values[5] + timedelta(days=31)).replace(day=1) - timedelta(days=1)),
                ((recap_values[5] + timedelta(days=31)).replace(day=1), (recap_values[5] + timedelta(days=62)).replace(day=1) - timedelta(days=1)),
                ((recap_values[5] + timedelta(days=62)).replace(day=1), recap_values[6])
            ]
            
        for idx, (start_date, end_date) in enumerate(date_ranges):
            date_list = pd.date_range(start=start_date, end=end_date).strftime('%d/%m/%Y').tolist()
            for ii, date in enumerate(date_list):
                worksheet.cell(row=17+ii, column=12+idx, value=date)
            
            if date_list:
                worksheet['C1'] = date_list[-1]
            else:
                print("La liste date_list est vide.")
                worksheet['C1'] = "Valeur par défaut"
        
    template_xlsx.save(path_rapport)
    template_xlsx.close()

# get timestamps (start and end)
def get_start_time(start_xlsx):
    timezone = "Indian/Antananarivo"
    s = pd.to_datetime(str(start_xlsx)).tz_localize(timezone)
    s = s.tz_convert(pytz.UTC)
    s = int(s.timestamp())
    return s

def get_end_time(end_xlsx):
    timezone = "Indian/Antananarivo"
    e = pd.to_datetime(str(end_xlsx)).tz_localize(timezone)
    e = e.tz_convert(pytz.UTC)
    e = int(e.timestamp()+3600*24-1)
    return e

def get_unix_timestamp(dt, timezone="Indian/Antananarivo", end_of_day=False):
    """
    Convert a given datetime to a Unix timestamp in UTC.

    :param dt: The datetime object or string to be converted.
    :param timezone: The local timezone of the input datetime.
    :param end_of_day: If True, adjust the timestamp to the end of the day.
    :return: Unix timestamp in UTC.
    """
    dt = pd.to_datetime(str(dt)).tz_localize(timezone).tz_convert(pytz.UTC)
    if end_of_day:
        dt += timedelta(hours=23, minutes=59, seconds=59)
    return int(dt.timestamp())

def write_data_site(writer, installation, start, end, rep_type=None):
    """
    Fetch and write various data for a given installation to an Excel file.

    :param writer: ExcelWriter object to write the data.
    :param installation: The installation object from which to fetch the data.
    :param start: The start timestamp.
    :param end: The end timestamp.
    :param rep_type: Report type identifier.
    """
    # Récupère toutes les données, en tenant compte des sources d'API et de CSV
    day_data_sun, day_data_conso, df_grouped_days, target_day_conso, target_day_sun = installation.get_all_data(start, end)
    day_data_sun, day_data_conso = installation.load_and_process_day_data()
    
    suffix = f"_{rep_type}" if rep_type else ""

    # Écriture des données de production et de consommation journalières dans Excel
    if day_data_sun is not None and day_data_conso is not None:
        day_data_sun.to_excel(writer, sheet_name=f"data_sun{suffix}", index=True)
        day_data_conso.to_excel(writer, sheet_name=f"data_conso{suffix}", index=True)
    else:
        print("Les données journalières day_data_sun ou day_data_conso sont manquantes pour cette installation.")
    
    # Écriture des données agrégées mensuelles dans Excel
    df_grouped_days.to_excel(writer, sheet_name=f"data{suffix}", index=True)

    # Gestion des données SOC selon le type d'installation
    if hasattr(installation, "api_endpoint"):  # Vérifie si l'installation utilise une API
        # Calcul des périodes pour les données SOC spécifiques
        s_sun, e_sun = int(target_day_sun.timestamp()), int(target_day_sun.timestamp()) + 86400
        s_conso, e_conso = int(target_day_conso.timestamp()), int(target_day_conso.timestamp()) + 86400

        # Appelle get_soc pour les installations avec API
        df_soc_sun = installation.get_soc(s_sun, e_sun)
        df_soc_conso = installation.get_soc(s_conso, e_conso)

        # Écriture des données SOC dans Excel
        df_soc_sun.to_excel(writer, sheet_name=f"SOC_sun{suffix}", index=False)
        df_soc_conso.to_excel(writer, sheet_name=f"SOC_conso{suffix}", index=False)
    else:
        print("L'installation ne supporte pas les données SOC ou ne dispose pas d'une API pour celles-ci.")


def fill_data(path_rapport, recap_values, installation):
    """
    Fill an Excel report with data for a given installation and time range.

    :param path_rapport: Path to the Excel report file.
    :param recap_values: Recap values containing start and end dates.
    :param installation: The installation object from which to fetch the data.
    """
    start, end = get_unix_timestamp(recap_values[5]), get_unix_timestamp(recap_values[6], end_of_day=True)
    s_prev_month_start = get_unix_timestamp((recap_values[5].replace(day=1) - timedelta(days=1)).replace(day=1))
    e_prev_month_end = get_unix_timestamp(recap_values[5].replace(day=1) - timedelta(days=1), end_of_day=True)
    
    with pd.ExcelWriter(path_rapport, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        if installation.report_type == '1m':
            write_data_site(writer, installation, start, end)
            bv_df, sy_df = installation.get_and_analyze_bv_and_sy(start, end)
            bv_df.to_excel(writer, sheet_name="Analysis Helper Battery", index=False)
            sy_df.to_excel(writer, sheet_name="Analysis Helper Solar Yield", index=False)
            
            prev_month_data = installation.get_data_previous_month(s_prev_month_start, e_prev_month_end)
            prev_month_data.to_excel(writer, sheet_name="data_previous_month", index=True)

        elif installation.report_type == '3m':
            # Define start and end times for the previous 3 months
            times = []
            for i in range(3):
                s_prev = get_unix_timestamp((recap_values[5] - timedelta(days=30 * (i + 1))).replace(day=1))
                e_prev = get_unix_timestamp(recap_values[5] - timedelta(days=30 * i), end_of_day=True)
                times.append((s_prev, e_prev))

            write_data_site(writer, installation, start, times[0][1], '1m')
            write_data_site(writer, installation, times[1][0], times[1][1], '2m')
            write_data_site(writer, installation, times[2][0], times[2][1], '3m')

            for i, (s_prev, e_prev) in enumerate(times):
                prev_month_data = installation.get_data_previous_month(s_prev, e_prev)
                prev_month_data.to_excel(writer, sheet_name=f"data_previous_month_{i + 1}m", index=True)

        s_12m = get_unix_timestamp(recap_values[5] - timedelta(days=365), end_of_day=True)
        data_12m = installation.get_data_12_months(s_12m, end)
        data_12m.to_excel(writer, sheet_name="data_12m", index=False)

def get_pv_gis_data(recap_values, report_file):

    print("Getting pv_gis data")
    # Initialize the parameters with defaults or values from the respective lists
    m_pv_gis = pd.Timestamp(recap_values[6]).month
    lat = recap_values[18]
    lon = recap_values[19]
    pvtechchoice = recap_values[20]

    # List of parameters to check for missing values and their default values
    params = [
        (recap_values[7], 0, "peakpower"),
        (recap_values[21], 0, "loss"),
        (recap_values[22], 0, "mountingplace"),
        (recap_values[23], 0, "angle"),
        (recap_values[24], 0, "azimut"),
        (recap_values[25], 0, "startyear"),
        (recap_values[26], 0, "endyear")
    ]

    # Dictionary to store the parameter values
    param_values = {}

    # Check for missing values and assign default values if necessary
    for value, default, name in params:
        if isinstance(value, str):
            param_values[name]=value
        else:
            param_values[name] = round(value) if not pd.isna(value) else default

    # Extract the parameter values from the dictionary
    peakpower = param_values["peakpower"]
    loss = param_values["loss"]
    mountingplace = param_values["mountingplace"]
    angle = param_values["angle"]
    azimut = param_values["azimut"]
    startyear = param_values["startyear"]
    endyear = param_values["endyear"]

    daily_mean_df, daily_mean_df_temp = get_irradiance_pv_gis(m_pv_gis, lat, lon, pvtechchoice, peakpower, loss, mountingplace, angle, azimut, startyear, endyear)
    with pd.ExcelWriter(report_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        daily_mean_df.to_excel(writer, sheet_name="Irradiance PVGIS", index=False)
            
        for i in range(1,6):
            if (m_pv_gis-i)%12 ==0:
                m_cond = 12
            else:
                m_cond = m_pv_gis-i
            daily_mean_df_parsed_prev = daily_mean_df_temp[daily_mean_df_temp['Month'] == m_cond]
        
            daily_mean_df_prev = pd.DataFrame()
            daily_mean_df_prev['Date'] = daily_mean_df_parsed_prev['Date']
            daily_mean_df_prev['Irradiance'] = daily_mean_df_parsed_prev['Irradiance']
            daily_mean_df_prev['Energie (Wh)'] = daily_mean_df_parsed_prev['Energie (Wh)']
            
            daily_mean_df_prev.to_excel(writer, sheet_name=f"Irradiance PVGIS prev_{i}m", index=False)
        
def get_alarm_data(report_file, recap_values, installation):
        start, end = get_unix_timestamp(recap_values[5]), get_unix_timestamp(recap_values[6], end_of_day=True)
        df_meta, df_alarm_summary, plot_files = installation.get_alarms(start, end)
        with pd.ExcelWriter(report_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

            if len(df_alarm_summary)==0:
                df_alarm_summary.to_excel(writer, sheet_name = "Alarm Summary", index = False)
            else:
                df_alarm_summary = pd.merge(df_alarm_summary, df_meta, left_on=0, right_on=0)
                df_alarm_summary = df_alarm_summary.sort_values(by=1)
                df_alarm_summary.to_excel(writer, sheet_name = "Alarm Summary", index = False)



def generate_report(master_path, template_path, result_path, installation):
    recap_values=get_recap_values(master_path, installation)
    report_file = create_report_file(installation, template_path, recap_values, result_path)
    print("Fichier du rapport cree")


    # Fill the report 
    fill_aide_rapport(report_file, recap_values, installation)
    print("Aide rapport terminée")
    fill_data(report_file, recap_values, installation)
    print("Data bien recupéré et introduit dans le rapport")
    if recap_values[3]=="Oui":
        get_pv_gis_data(recap_values, report_file)
        print("Données PV_Gis bien récupérés")
    if recap_values[2] == "Oui":
        get_alarm_data(report_file, recap_values, installation)
        print("DOnnées des alarmes bien récupérées et traitées")
    
    create_charts(report_file, recap_values[4])
    print("Graphiques générés")
    print("Rapport terminé")

