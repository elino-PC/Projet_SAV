from openpyxl import load_workbook
from openpyxl.chart import  BarChart, PieChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font 
from openpyxl.drawing.image import Image  

def create_charts(report_file, period):
    workbook = load_workbook(report_file)
    sheet_help = workbook["Aide Rapport"]    
    sheet_report = workbook["Rapport"]
# Create pie chart            
    chart = PieChart()
    chart.title = "Provenance de l'énergie"
    font = Font(typeface='Times New Roman')
    cp = CharacterProperties(latin=font, sz=1600, b=True, solidFill='484848')
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)]) 
    chart.title.txPr = rtp
    chart.title.tx.rich.p[0].pPr = pp
    data = Reference(sheet_help, min_col=2, min_row=11, max_row=14, max_col=2)
    categories = Reference(sheet_help, min_col=1, min_row=12, max_row=14)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList() 
    chart.dataLabels.showVal = True
    cp_ticks = CharacterProperties(latin=font, sz=1000, b=True, solidFill='ffffff')
    chart.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_ticks), endParaRPr=cp_ticks)])
    chart.width = 15.25
    
    colors = ['00b050', '9bbb59', '953735']
    series = chart.series[0]
    for jc in range(len(colors)):
        pt = DataPoint(idx=jc)
        pt.graphicalProperties.solidFill = colors[jc]
        series.dPt.append(pt)
        
    sheet_report.add_chart(chart, "C84")
    
                            
    # Create a bar chart
    if period =='1 mois':
        bar_chart = BarChart()
        sheet_data = workbook["data"] 
        bar_chart.title = "Provenance de l'énergie consommée par jour"
        font = Font(typeface='Times New Roman')
        cp = CharacterProperties(latin=font, sz=1600, b=False, solidFill='484848')
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        bar_chart.title.txPr = rtp
        bar_chart.title.tx.rich.p[0].pPr = pp
        bar_data = Reference(sheet_data, min_col=18, min_row=1, max_row=32, max_col=21)
        titles = Reference(sheet_help, min_col=1, min_row=17, max_row=47, max_col=1)
        
        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.grouping = "stacked"
        bar_chart.overlap = 100
        bar_chart.x_axis.title = "Jour"
        bar_chart.y_axis.title = "Energie en kWh"
        bar_chart.width = 17.76
        bar_chart.height = 7.25
        bar_chart.set_categories(titles)
        
        cp_text = CharacterProperties(latin=font, sz=900, b=True, solidFill='484848')
        cp_axis = CharacterProperties(latin=font, sz=900, b=False, solidFill='484848')
        bar_chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
        bar_chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
        
        pp = ParagraphProperties(defRPr=cp_text)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp_text)]) 
        bar_chart.x_axis.title.txPr = rtp
        bar_chart.x_axis.title.tx.rich.p[0].pPr = pp
        bar_chart.x_axis.txPr.properties.rot = "-2700000"
        bar_chart.y_axis.title.txPr = rtp
        bar_chart.y_axis.title.tx.rich.p[0].pPr = pp
        
        colors = ["953735", "0070c0", "0fb73d", "953735"]
        series = bar_chart.series
        for jc in range(len(colors)):
            series[jc].graphicalProperties.solidFill = colors[jc]
        
        sheet_report.add_chart(bar_chart, "A101")
    else:
        for i_bc in range(1,4):
            bar_chart = BarChart()
            sheet_data = workbook[f"data_{i_bc}m"]
            bar_chart.title = "Provenance de l'énergie consommée par jour"
            font = Font(typeface='Times New Roman')
            cp = CharacterProperties(latin=font, sz=1600, b=False, solidFill='484848')
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            bar_chart.title.txPr = rtp
            bar_chart.title.tx.rich.p[0].pPr = pp
            bar_data = Reference(sheet_data, min_col=18, min_row=1, max_row=32, max_col=21)
            titles = Reference(sheet_help, min_col=11+i_bc, min_row=17, max_row=47, max_col=11+i_bc)
            
            bar_chart.add_data(bar_data, titles_from_data=True)
            bar_chart.grouping = "stacked"
            bar_chart.overlap = 100
            bar_chart.x_axis.title = "Jour"
            bar_chart.y_axis.title = "Energie en kWh"
            bar_chart.width = 17.76
            bar_chart.height = 7.25
            bar_chart.set_categories(titles)
            
            cp_text = CharacterProperties(latin=font, sz=900, b=True, solidFill='484848')
            cp_axis = CharacterProperties(latin=font, sz=900, b=False, solidFill='484848')
            bar_chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
            bar_chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
            
            pp = ParagraphProperties(defRPr=cp_text)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp_text)]) 
            bar_chart.x_axis.title.txPr = rtp
            bar_chart.x_axis.title.tx.rich.p[0].pPr = pp
            bar_chart.x_axis.txPr.properties.rot = "-2700000"
            bar_chart.y_axis.title.txPr = rtp
            bar_chart.y_axis.title.tx.rich.p[0].pPr = pp
            
            colors = ["953735", "0070c0", "0fb73d", "953735"]
            series = bar_chart.series
            for jc in range(len(colors)):
                series[jc].graphicalProperties.solidFill = colors[jc]
            cell_pos = 101+(i_bc-1)*50
            sheet_report.add_chart(bar_chart, f"A{cell_pos}")
                
    # Create a line/bar chart day sun
    combo_chart_sun = BarChart()
    sheet_data = workbook["data_sun"] 
    date_title = sheet_data['A2'].value
    date_title = date_title.strftime("%d/%m/%Y")
    title_combo_sun = f"Journée de production solaire maximale ({date_title})"
    combo_chart_sun.title = title_combo_sun
    font = Font(typeface='Times New Roman')
    cp = CharacterProperties(latin=font, sz=1600, b=True, solidFill='1f497d')
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)]) 
    combo_chart_sun.title.txPr = rtp
    combo_chart_sun.title.tx.rich.p[0].pPr = pp
    
    bar_data = Reference(sheet_data, min_col=22, min_row=1, max_row=25, max_col=27)
    titles = Reference(sheet_help, min_col=3, min_row=16, max_row=39, max_col=3)
    combo_chart_sun.add_data(bar_data, titles_from_data=True)
    combo_chart_sun.grouping = "stacked"
    combo_chart_sun.overlap = 100
    combo_chart_sun.x_axis.title = "Heure"
    combo_chart_sun.y_axis.title = "Puissance en kW"
    
    cp_axis = CharacterProperties(latin=font, sz=900, b=False, solidFill='1f497d')
    combo_chart_sun.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    combo_chart_sun.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    cp_text = CharacterProperties(latin=font, sz=900, b=True, solidFill='1f497d')
    pp = ParagraphProperties(defRPr=cp_text)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp_text)]) 
    combo_chart_sun.x_axis.title.txPr = rtp
    combo_chart_sun.x_axis.title.tx.rich.p[0].pPr = pp
    combo_chart_sun.y_axis.title.txPr = rtp
    combo_chart_sun.y_axis.title.tx.rich.p[0].pPr = pp
    
    combo_chart_sun.width = 17.76
    combo_chart_sun.height = 9.75
    combo_chart_sun.legend.position = 'b'
    combo_chart_sun.set_categories(titles)
    colors = ["f7cf3b", "70ad47", "a5a5a5", "ed7d31", "0070c0" ,"953735"]
    series = combo_chart_sun.series
    for jc in range(len(colors)):
        series[jc].graphicalProperties.solidFill = colors[jc]
    
    line_chart = LineChart()
    line_data = Reference(sheet_data, min_col=28, min_row=1, max_row=25, max_col=29)
    line_chart.add_data(line_data, titles_from_data=True)
    colors = ["f7cf3b", "0070c0"]
    series = line_chart.series
    for jc in range(len(colors)):
        series[jc].graphicalProperties.line.solidFill = colors[jc]
    combo_chart_sun += line_chart
    
    line_chart_soc = LineChart()
    line_data = Reference(sheet_help, min_col=5, min_row=15, max_row=39, max_col=5)
    line_chart_soc.add_data(line_data, titles_from_data=True)
    line_chart_soc.y_axis.axId = 200
    line_chart_soc.y_axis.scaling.max  = 1
    line_chart_soc.y_axis.scaling.min  = 0
    line_chart_soc.y_axis.title = "Etat de charge de la batterie"
    line_chart_soc.y_axis.crosses = "max"
    line_chart_soc.y_axis.number_format = '0%'  
    line_chart_soc.y_axis.majorGridlines = None
    line_chart_soc.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    line_chart_soc.y_axis.title.txPr = rtp
    line_chart_soc.y_axis.title.tx.rich.p[0].pPr = pp
    
    series = line_chart_soc.series[0]
    series.graphicalProperties.line.dashStyle = "sysDot"
    colors = "1f497d"
    series.graphicalProperties.line.solidFill = colors
    
    combo_chart_sun += line_chart_soc
    if period=='1 mois':
        sheet_report.add_chart(combo_chart_sun, "A153")
    else:
        sheet_report.add_chart(combo_chart_sun, "A253")
    
    # Create a line/bar chart day conso
    combo_chart_conso = BarChart()
    sheet_data = workbook["data_conso"] 
    date_title = sheet_data['A2'].value
    date_title = date_title.strftime("%d/%m/%Y")
    title_combo_conso = f"Journée de consommation maximale ({date_title})"
    combo_chart_conso.title = title_combo_conso
    font = Font(typeface='Times New Roman')
    cp = CharacterProperties(latin=font, sz=1600, b=True, solidFill='1f497d')
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)]) 
    combo_chart_conso.title.txPr = rtp
    combo_chart_conso.title.tx.rich.p[0].pPr = pp
    
    bar_data = Reference(sheet_data, min_col=22, min_row=1, max_row=25, max_col=27)
    titles = Reference(sheet_help, min_col=3, min_row=16, max_row=39, max_col=3)
    combo_chart_conso.add_data(bar_data, titles_from_data=True)
    combo_chart_conso.grouping = "stacked"
    combo_chart_conso.overlap = 100
    combo_chart_conso.x_axis.title = "Heure"
    combo_chart_conso.y_axis.title = "Puissance en kW"
    
    cp_axis = CharacterProperties(latin=font, sz=900, b=False, solidFill='1f497d')
    combo_chart_conso.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    combo_chart_conso.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    cp_text = CharacterProperties(latin=font, sz=900, b=True, solidFill='1f497d')
    pp = ParagraphProperties(defRPr=cp_text)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp_text)]) 
    combo_chart_conso.x_axis.title.txPr = rtp
    combo_chart_conso.x_axis.title.tx.rich.p[0].pPr = pp
    combo_chart_conso.y_axis.title.txPr = rtp
    combo_chart_conso.y_axis.title.tx.rich.p[0].pPr = pp
    
    combo_chart_conso.width = 17.76
    combo_chart_conso.height = 9
    combo_chart_conso.legend.position = 'b'
    combo_chart_conso.set_categories(titles)
    colors = ["f7cf3b", "70ad47", "a5a5a5", "ed7d31", "0070c0" ,"953735"]
    series = combo_chart_conso.series
    for jc in range(len(colors)):
        series[jc].graphicalProperties.solidFill = colors[jc]
    
    line_chart = LineChart()
    line_data = Reference(sheet_data, min_col=28, min_row=1, max_row=25, max_col=29)
    line_chart.add_data(line_data, titles_from_data=True)
    colors = ["f7cf3b", "0070c0"]
    series = line_chart.series
    for jc in range(len(colors)):
        series[jc].graphicalProperties.line.solidFill = colors[jc]
    combo_chart_conso += line_chart
    
    line_chart_soc = LineChart()
    line_data = Reference(sheet_help, min_col=6, min_row=15, max_row=39, max_col=6)
    line_chart_soc.add_data(line_data, titles_from_data=True)
    line_chart_soc.y_axis.axId = 200
    line_chart_soc.y_axis.scaling.max  = 1
    line_chart_soc.y_axis.scaling.min  = 0
    line_chart_soc.y_axis.title = "Etat de charge de la batterie"
    line_chart_soc.y_axis.crosses = "max"
    line_chart_soc.y_axis.number_format = '0%'  
    line_chart_soc.y_axis.majorGridlines = None
    line_chart_soc.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    line_chart_soc.y_axis.title.txPr = rtp
    line_chart_soc.y_axis.title.tx.rich.p[0].pPr = pp
    
    series = line_chart_soc.series[0]
    series.graphicalProperties.line.dashStyle = "sysDot"
    colors = "1f497d"
    series.graphicalProperties.line.solidFill = colors
    
    combo_chart_conso += line_chart_soc
        
    if period=='1 mois':
        sheet_report.add_chart(combo_chart_conso, "A176")
    else:
        sheet_report.add_chart(combo_chart_conso, "A276")
    
    # Create a bar chart (12 month)
    combo_chart = BarChart()
    sheet_data = workbook["data_12m"] 
    bar_data = Reference(sheet_data, min_col=2, min_row=1, max_row=14, max_col=5)
    titles = Reference(sheet_help, min_col=2, min_row=17, max_row=29, max_col=2)
    combo_chart.add_data(bar_data, titles_from_data=True)
    combo_chart.type = "col"
    combo_chart.grouping = "stacked"
    combo_chart.overlap = 100
    combo_chart.x_axis.title = "Mois"
    combo_chart.y_axis.title = "Energie en kWh"
    combo_chart.set_categories(titles)
    combo_chart.width = 17.76
    combo_chart.height = 9.5
    combo_chart.legend.position = 'b'
    
    cp_axis = CharacterProperties(latin=font, sz=900, b=False, solidFill='1f497d')
    combo_chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    combo_chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    cp_text = CharacterProperties(latin=font, sz=1200, b=True, solidFill='1f497d')
    pp = ParagraphProperties(defRPr=cp_text)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp_text)]) 
    combo_chart.x_axis.title.txPr = rtp
    combo_chart.x_axis.title.tx.rich.p[0].pPr = pp
    combo_chart.y_axis.title.txPr = rtp
    combo_chart.y_axis.title.tx.rich.p[0].pPr = pp
    combo_chart.x_axis.txPr.properties.rot = "-2700000"
    
    colors = ["f7cf3b","c07000","c04a00", "ed7d31"]
    series = combo_chart.series
    series_length=[]
    for col in sheet_data.iter_cols(min_col=1, values_only=True):
        series_length.append(len(col))

    for jc in range(len(colors)):
        series[jc].graphicalProperties.solidFill = colors[jc]
        
    combo_chart.title = f"Historique sur les {series_length[0]-1} derniers mois"
    font = Font(typeface='Calibri')
    cp = CharacterProperties(latin=font, sz=1600, b=False, solidFill='484848')
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)]) 
    combo_chart.title.txPr = rtp
    combo_chart.title.tx.rich.p[0].pPr = pp

    if period=='1 mois':
        sheet_report.add_chart(combo_chart, "A202")
    else:
        sheet_report.add_chart(combo_chart, "A302")
    workbook.save(report_file)
    print("done with charts")