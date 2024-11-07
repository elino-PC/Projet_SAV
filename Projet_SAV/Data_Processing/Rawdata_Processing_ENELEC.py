# Mon premier script de traitement des donn√©es ENELEC

import openpyxl
from time import *
from datetime import datetime, timedelta
import os
import pandas as pd
import shutil
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Constantes pour les noms de fichiers
PREFIX_RAWDATA = "RAWDATA_"
PREFIX_TRAITE = "traite_"
EXTENSION_CSV = ".csv"
EXTENSION_XLSX = ".xlsx"

def traiter_et_regrouper_rawdata(raw_data_in_folder, alldatas_out_folder):
    os.makedirs(alldatas_out_folder, exist_ok=True)
    
    files_to_process = [f for f in os.listdir(raw_data_in_folder) if f.startswith(PREFIX_RAWDATA) and f.endswith(EXTENSION_CSV)]

    all_dfs = []
    column_names_dict = {}  # Dictionnaire pour stocker les noms de colonnes d'origine

    for raw_data_in_file in files_to_process:
        print(f"Lecture du fichier : {raw_data_in_file}")
        try:
            # Lire le fichier CSV avec les caractéristiques spécifiques
            df = pd.read_csv(os.path.join(raw_data_in_folder, raw_data_in_file), sep=';', decimal=',', encoding='ISO-8859-1', skiprows=1)
            print("Lecture réussie.")

            # Vérifier les doublons
            duplicate_rows = df[df.duplicated()]
            if not duplicate_rows.empty:
                # Supprimer les doublons
                df = df.drop_duplicates(subset=df.columns)
                print(f"{len(duplicate_rows)} lignes doublons supprimées.")
            else:
                print("Aucune ligne doublon trouvée.")
            
            # Supprimer les lignes qui ne sont pas des valeurs dans les lignes 3 à 1442
            df = df.dropna(subset=df.columns[2:1442], how='all')
            df = df[~df.iloc[:, 0].str.contains(' TimeStamp ', na=False)]
            
            # Limiter le DataFrame à 1442 lignes
            df = df.head(1442)

            # Remplacer les points par des virgules
            df = df.apply(lambda x: x.map(lambda x: str(x).replace('.', ',') if '.' in str(x) else x))            
            
            # Convertir les colonnes appropriées en types de données numériques
            for col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col].apply(lambda x: str(x).lstrip("'").replace(',', '.') if isinstance(x, str) else x), errors='raise')
                except ValueError:
                    # Si la conversion échoue, la colonne reste inchangée
                    pass

            # Convertir la colonne 'TimeStamp' en type datetime et en faire l'index
            df[' TimeStamp '] = pd.to_datetime(df[' TimeStamp '])
            df.set_index(' TimeStamp ', inplace=True)  # Définir 'TimeStamp' comme index
            
            # Réindexer le DataFrame pour inclure toutes les dates et heures entre le premier et le dernier enregistrement
            min_date = df.index.min()
            max_date = df.index.max()
            idx = pd.date_range(start=min_date, end=max_date, freq='min')
            df = df.reindex(idx)
            
            df.fillna(0, inplace=True)
            
            # Créer une feuille de calcul pour chaque fichier traité
            all_dfs.append((raw_data_in_file, df))
            column_names_dict[raw_data_in_file] = df.columns.tolist()  # Enregistrer les noms de colonnes

            if not df.empty:
                # Déplacer le fichier traité vers le dossier de sortie
                shutil.copy(os.path.join(raw_data_in_folder, raw_data_in_file), os.path.join(alldatas_out_folder, f"{PREFIX_TRAITE}{raw_data_in_file}"))

                print(f"Fichier {raw_data_in_file} traité et copié avec succès.")
            else:
                print(f"Avertissement : Aucune donnée dans le fichier {raw_data_in_file}")

        except Exception as e:
            print(f"Erreur lors de la lecture du fichier {raw_data_in_file} : {e}")

    alldatas_excel_path = os.path.join(alldatas_out_folder, "Alldatas_202410.xlsx")

    # Créer un gestionnaire de fichiers Excel
    with pd.ExcelWriter(alldatas_excel_path, engine='openpyxl') as writer:
        
        # Trier les feuilles par date avant de les écrire dans le fichier Excel
        all_dfs.sort(key=lambda x: datetime.strptime(x[0][8:-len(EXTENSION_CSV)], '%Y-%m-%d'))
        for raw_data_in_file, df in all_dfs:
            sheet_name = os.path.splitext(raw_data_in_file)[0] 
            
            # Ajouter la feuille de calcul au classeur Excel
            df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=2, header=False)
            
            # Ajouter le titre au-dessus du tableau
            worksheet = writer.sheets[sheet_name]       
            worksheet.cell(row=2, column=1, value='TimeStamp')
            worksheet['A1'] = f"Fichier de données journalières, journée du {raw_data_in_file[8:-len(EXTENSION_CSV)]}"
            
            # Ajouter les titres des colonnes à partir de 'B2'
            for col_num, col_label in enumerate(df.columns, start=2):
                cell = worksheet.cell(row=2, column=col_num, value=col_label)
                
            # Remplir les cellules vides par 0
            for row in worksheet.iter_rows(min_row=3, min_col=1, max_row=len(df) + 2, max_col=len(df.columns)):
                for cell in row:
                    if cell.value is None:
                        cell.value = 0
            
            # Accéder à la feuille de calcul pour ajouter le tableau
            worksheet.sheet_state = 'visible'

    print(f"Le classeur Excel a été enregistré dans {alldatas_excel_path}")
    print("Opération terminée.")


# Exemple d'utilisation
raw_data_in_folder = "G:/.shortcut-targets-by-id/1FQfz_wNk7M-PeQeUyVAHoy9ay4UY1_62/3 - SAV/3 - Rapports de production/3 - Outil de rapport/Projet_SAV/Projet_SAV/Data_Processing/Data_Folders/"
alldatas_out_folder = "G:/.shortcut-targets-by-id/1FQfz_wNk7M-PeQeUyVAHoy9ay4UY1_62/3 - SAV/3 - Rapports de production/3 - Outil de rapport/Projet_SAV/Rapports de production/"
 
traiter_et_regrouper_rawdata(raw_data_in_folder, alldatas_out_folder)